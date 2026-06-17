#!/usr/bin/env python3
"""신문산업_매출액_*.xlsx → table-data.json (브라우저 표 렌더링·AI 컨텍스트용)."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = next(ROOT.glob("신문산업_매출액_*.xlsx"), None)
OUT = ROOT / "table-data.json"


def cell_val(ws, r: int, c: int):
    v = ws.cell(r, c).value
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def resolve_cell(ws, anchor: dict[tuple[int, int], tuple[int, int, object]], r: int, c: int):
    key = (r, c)
    if key in anchor:
        _, _, val = anchor[key]
        return val
    return cell_val(ws, r, c)


def build_anchor_map(ws) -> dict[tuple[int, int], tuple[int, int, object]]:
    anchor: dict[tuple[int, int], tuple[int, int, object]] = {}
    for m in ws.merged_cells.ranges:
        min_r, min_c = m.min_row, m.min_col
        val = cell_val(ws, min_r, min_c)
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                anchor[(r, c)] = (min_r, min_c, val)
    return anchor


def build_columns(ws, anchor) -> list[dict]:
    cols: list[dict] = []
    year = ""
    for c in range(4, ws.max_column + 1):
        y = resolve_cell(ws, anchor, 1, c)
        if y != "":
            year = str(y)
        metric = str(resolve_cell(ws, anchor, 2, c))
        cols.append(
            {
                "year": year,
                "metric": metric,
                "label": f"{year} · {metric}",
            }
        )
    return cols


def build_records(ws, anchor, columns) -> list[dict]:
    records: list[dict] = []
    cat = ["", "", ""]

    for r in range(3, ws.max_row + 1):
        for i in range(3):
            v = resolve_cell(ws, anchor, r, i + 1)
            if v != "":
                cat[i] = str(v)
                for j in range(i + 1, 3):
                    cat[j] = ""

        item = " / ".join(x for x in cat if x)
        if not item:
            continue

        for i, col in enumerate(columns):
            val = resolve_cell(ws, anchor, r, i + 4)
            if val == "" or val is None or val == "-":
                continue
            records.append(
                {
                    "item": item,
                    "year": col["year"],
                    "metric": col["metric"],
                    "value": val,
                }
            )
    return records


def build_rows(ws, anchor) -> list[list[dict]]:
    merged: dict[tuple[int, int], dict] = {}
    for m in ws.merged_cells.ranges:
        min_r, min_c = m.min_row, m.min_col
        max_r, max_c = m.max_row, m.max_col
        val = cell_val(ws, min_r, min_c)
        span_r, span_c = max_r - min_r + 1, max_c - min_c + 1
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                merged[(r, c)] = {
                    "skip": not (r == min_r and c == min_c),
                    "v": val,
                    "rowspan": span_r,
                    "colspan": span_c,
                }

    rows: list[list[dict]] = []
    for r in range(1, ws.max_row + 1):
        row: list[dict] = []
        for c in range(1, ws.max_column + 1):
            key = (r, c)
            if key in merged:
                m = merged[key]
                if m["skip"]:
                    row.append({"skip": True})
                else:
                    cell: dict = {"v": m["v"]}
                    if m["rowspan"] > 1:
                        cell["rowspan"] = m["rowspan"]
                    if m["colspan"] > 1:
                        cell["colspan"] = m["colspan"]
                    row.append(cell)
            else:
                row.append({"v": cell_val(ws, r, c)})
        rows.append(row)
    return rows


def main() -> None:
    if not XLSX or not XLSX.is_file():
        raise SystemExit(f"엑셀 파일이 없습니다: {ROOT / '신문산업_매출액_*.xlsx'}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    anchor = build_anchor_map(ws)
    columns = build_columns(ws, anchor)
    records = build_records(ws, anchor, columns)
    rows = build_rows(ws, anchor)

    payload = {
        "title": "신문산업 매출액",
        "source": XLSX.name,
        "schema": {
            "description": "한국언론진흥재단 KOSIS 신문·잡지산업실태조사 — 신문산업 매출액",
            "units": "매출액·기업공시업체매출액 단위=백만원, 기업공시업체수 단위=개",
            "itemHierarchy": "신문산업별(1) > 신문산업별(2) > 신문산업별(3) 계층으로 항목 구분",
            "years": "2010–2024 (2012–2017년은 연도별 매출액·기업공시업체수·기업공시업체매출액 3열)",
        },
        "columns": columns,
        "records": records,
        "rows": rows,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT.name} ({len(rows)} rows, {len(records)} records) from {XLSX.name}")


if __name__ == "__main__":
    main()
