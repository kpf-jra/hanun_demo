#!/usr/bin/env python3
"""kolang_2nd.xlsx → bank.json 및 index.html 내 JSON 갱신.

실행 예:
  cd demos/kolang-quiz && python build-bank.py
  npm run build:kolang-quiz   # 저장소 루트에서
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "kolang_2nd.xlsx"
BANK_JSON = ROOT / "bank.json"
INDEX_HTML = ROOT / "index.html"


def split_phrases(value: object) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split("/") if part.strip()]


def load_bank() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    bank: list[dict] = []

    ws = wb["자꾸 틀리는 표현"]
    for row in list(ws.iter_rows(values_only=True))[1:]:
        wrong, right, explain = row[0], row[1], row[2]
        if not wrong or not right:
            continue
        bank.append(
            {
                "type": "mistake",
                "wrong": split_phrases(wrong),
                "right": split_phrases(right),
                "explain": str(explain).strip() if explain else "",
            }
        )

    ws = wb["겹말"]
    for row in list(ws.iter_rows(values_only=True))[1:]:
        wrong, right = row[0], row[1]
        if not wrong or not right:
            continue
        bank.append(
            {
                "type": "overlap",
                "wrong": split_phrases(wrong),
                "right": split_phrases(right),
            }
        )

    return bank


def patch_index_html(bank: list[dict]) -> None:
    html = INDEX_HTML.read_text(encoding="utf-8")
    payload = json.dumps(bank, ensure_ascii=False, separators=(",", ":"))
    pattern = r'(<script type="application/json" id="uc-bank-json">\s*)\[.*?\](\s*</script>)'
    replacement = rf"\1{payload}\2"
    updated, count = re.subn(pattern, replacement, html, count=1, flags=re.DOTALL)
    if count != 1:
        raise RuntimeError("index.html에서 uc-bank-json 블록을 찾지 못했습니다.")
    INDEX_HTML.write_text(updated, encoding="utf-8")


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"엑셀 파일이 없습니다: {XLSX}")

    bank = load_bank()
    BANK_JSON.write_text(
        json.dumps(bank, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    patch_index_html(bank)

    mistake = sum(1 for item in bank if item["type"] == "mistake")
    overlap = sum(1 for item in bank if item["type"] == "overlap")
    print(f"bank.json 갱신: 총 {len(bank)}문항 (자꾸 틀리는 표현 {mistake}, 겹말 {overlap})")
    print(f"index.html 내 JSON도 갱신했습니다.")


if __name__ == "__main__":
    main()
